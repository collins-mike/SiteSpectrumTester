'''
Drone race site spectrum tester
Created on Nov 2, 2017 NextGen RF design

@author: Mike Collins
'''
#import libraries
import sys, os, random,csv,time
from PyQt4.QtCore import *
from PyQt4.QtGui import *

import multiprocessing,logging

import matplotlib
from matplotlib.backends.backend_qt4agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt4agg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure
from mpl_toolkits.mplot3d import Axes3D
from matplotlib import cm
from matplotlib import ticker

import openpyxl as pyxl
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.drawing.image import Image
from openpyxl import load_workbook

#import files
from SignalHound import *
from specan import *
from PyQt4.Qt import QTextEdit
from advSettingDialog.advSettingDialog import advSettingsDialog
import atexit
from SpecTest import SpecTest



class Application(QMainWindow):
    def __init__(self, parent=None):
        QMainWindow.__init__(self, parent)
        self.setWindowTitle("Race Site RF Spectrum Tester")
        self.createForm()
        self.foundSpec=False
        self.specan=SpecAnalyzer()
        self.PROG=0
        
        self.plotImageList=[]
        
        self.settings=advSettingsDialog(self)
        
    def createForm(self):
        #=======================================================================
        #
        #          Name:    createForm
        #
        #    Parameters:    none
        #
        #        Return:    none
        #
        #   Description:    create GUI
        #
        #=======================================================================
        
        #=======================================================================
        # setup layout
        #=======================================================================
        self.main_frame=QWidget()
        vbox=QVBoxLayout()
        fbox=QFormLayout()
        vbox.addLayout(fbox)
        self.main_frame.setLayout(vbox)
        
        self.btn_findDevice=QPushButton("Find Spectrum Analyzer")
        self.btn_findDevice.setFixedHeight(50)
        self.btn_findDevice.setFixedWidth(150)
        self.btn_findDevice.setEnabled(True)
        self.connect(self.btn_findDevice, SIGNAL('clicked()'), self.click_find)
        self.deviceInfo=QLineEdit("No spectrum analyzer found")
        self.deviceInfo.setEnabled(False)
        fbox.addRow(self.btn_findDevice,self.deviceInfo)
    
        self.btn_run=QPushButton("Run")
        self.btn_run.setFixedHeight(50)
        self.btn_run.setFixedWidth(150)
        self.btn_run.setEnabled(False)
        self.connect(self.btn_run, SIGNAL('clicked()'), self.click_Run)
        self.runInfo=QLineEdit("Find Spectrum analyzer before running test")
        self.runInfo.setEnabled(False)
        fbox.addRow(self.btn_run,self.runInfo)
        
        self.btn_saveAs=QPushButton("Save As")
        self.btn_saveAs.setFixedHeight(50)
        self.btn_saveAs.setFixedWidth(150)
        self.btn_saveAs.setEnabled(False)
        self.connect(self.btn_saveAs, SIGNAL('clicked()'), self.saveData)
        self.saveInfo=QLineEdit('No Data to Save')
        self.saveInfo.setEnabled(False)
        fbox.addRow(self.btn_saveAs,self.saveInfo)
        
        
        self.btn_advSettings=QPushButton("Advanced Settings")
        self.btn_advSettings.setFixedHeight(50)
        self.btn_advSettings.setFixedWidth(150)
        self.btn_advSettings.setEnabled(True)
        self.connect(self.btn_advSettings, SIGNAL('clicked()'), self.click_advSettings)
        hbox=QHBoxLayout()
        hbox.addStretch()
        hbox.addWidget(self.btn_advSettings)
        hbox.addStretch()
        
        #=======================================================================
        # create plot
        #=======================================================================
        self.fig = Figure(figsize=(12,5))
        self.canvas = FigureCanvas(self.fig)
        
        self.canvas.setParent(self.main_frame)
        self.plot = self.fig.add_subplot(111)
        self.plot.set_title('---',fontsize=14,fontweight=200)       
        self.plot.set_ylim([-100,-50])
        
        vbox.addWidget(self.canvas)
        vbox.addLayout(hbox)
        self.setCentralWidget(self.main_frame)
        
        #=======================================================================
        # create standard tests
        #=======================================================================
        self.buildDefaultTests()

    def buildDefaultTests(self):
        self.testList=[]
        self.testList.append(SpecTest(parent=self, plot=self.plot, testNum=0, name="5.5 GHz",rbw=157.1e3,sweepTime=0.01,sweepNum=20,freqCenter=5500e6,freqSpan=1e9,threshold=-50))
        self.testList.append(SpecTest(parent=self, plot=self.plot, testNum=1, name="4 GHz",rbw=157.1e3,sweepTime=0.01,sweepNum=20,freqCenter=4000e6,freqSpan=1e9,threshold=-50))
        self.testList.append(SpecTest(parent=self, plot=self.plot, testNum=2, name="915 MHz",rbw=39.45e3,sweepTime=0.01,sweepNum=20,freqCenter=915e6,freqSpan=100e6,threshold=-50))
        self.testList.append(SpecTest(parent=self, plot=self.plot, testNum=3, name="863 MHz",rbw=39.45e3,sweepTime=0.01,sweepNum=20,freqCenter=863e6,freqSpan=100e6,threshold=-50))
        self.testList.append(SpecTest(parent=self, plot=self.plot, testNum=4, name="Wide Band",rbw=315.6e3,sweepTime=0.001,sweepNum=20,freqCenter=3015e6,freqSpan=5970e6,threshold=-50))
        
    def click_Run(self):
        #=======================================================================
        #
        #          Name:    click_Run
        #
        #    Parameters:    none
        #
        #        Return:    none
        #
        #   Description:    runs consecutive sweeps in 5 different bandwidths
        #
        #=======================================================================
        
        self.purgeplotImageList()
        
        canceled=False
        
        self.btn_run.setEnabled(False)
        self.runInfo.setText('Running Test...')
        self.btn_saveAs.setEnabled(False)
        
        self.updateUi()
        
        totalSweeps=0
        for i in range(len(self.testList)):
            totalSweeps+=self.testList[i].sweepNum
        
        self.progress=QProgressDialog(labelText="Running Test",minimum=0,maximum=totalSweeps)
        PROG=0
        self.progress.setWindowTitle('Test Progress')
        self.progress.show()

        for i in range(len(self.testList)):
            canceled=self.testList[i].runSweep()
            if canceled:
                break
            
        #=======================================================================
        # Test Complete
        #=======================================================================
        self.progress.close()
        
        msg = QMessageBox()
        msg.setWindowTitle('Test Complete')
        msg.setInformativeText("Test complete!\n\nWould you like to save this data?")
        msg.setIcon(QMessageBox.Information)
        msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        msg.buttonClicked.connect(self.msgbtn)
        
        if canceled==False:
            msg.exec_()
            self.btn_saveAs.setEnabled(True)
            self.saveInfo.setText('Ready to save data from last test')
        else:
            self.btn_saveAs.setEnabled(False)  
            self.saveInfo.setText('Test canceled, No Data to test')
              
        self.runInfo.setText('Ready to run test')
        self.btn_run.setEnabled(True)
        self.PROG=0
        
    def msgbtn(self,i):
        #=======================================================================
        #
        #          Name:    msgbtn
        #
        #    Parameters:    i is an instance of a yes or no question dialog
        #
        #        Return:    none    
        #
        #   Description:    sets behavior after yes or no is clicked on by user
        #
        #=======================================================================
        print "Button pressed is:",i.text()
        if i.text()== "&Yes":
            self.saveData()
            
        elif i.text()== "&No":
            pass    
    
    def updateUi(self):
            QApplication.instance().processEvents()
    
    def saveData(self):
        #=======================================================================
        #
        #          Name:    saveData
        #
        #    Parameters:    None    
        #
        #        Return:    None    
        #
        #   Description:    saves all plots to a .xlsx file, give warning message if save error
        #
        #=======================================================================
        
        self.btn_saveAs.setEnabled(False)
        self.btn_run.setEnabled(False)
        self.updateUi()
        
        file_choices = "Excel Workbook ( *.xlsx)"
        path = unicode(QFileDialog.getSaveFileName(self, 'Save', '', file_choices))
        
        
        
        if path:
            wb = pyxl.Workbook()
            ws = wb.create_sheet("Race Site Spectrum Test", 0)
            
            i = 0
            for pltImg in self.plotImageList:
                
                img = Image(pltImg)
                ws.add_image(img, 'A'+str(i*25+1))
                i+=1
            
        try:
            wb.save(path)
        except: 
            print "file save Error"
            self.show_errorDialog("File Save Error!", "Unable to save report!", "Ensure save location is not open in another program.")
                  
        self.btn_saveAs.setEnabled(True)
        self.btn_run.setEnabled(True)

    
    def click_advSettings(self):
        self.settings.disableTableModification=True
        self.settings.updateTable()
        self.settings.disableTableModification=False
        self.settings.exec_()
        
    def purgeplotImageList(self):
        for filename in self.plotImageList:
            if os.path.exists(filename):
                os.remove(filename)
            else:
                print("Sorry, I can not remove %s file." % filename)
        
        self.plotImageList=[]
        
    def click_find(self):
        #=======================================================================
        #
        #          Name:    click_find
        #
        #    Parameters:    none    
        #
        #        Return:    none
        #
        #   Description:    searches for specan and displays result in GUI
        #
        #=======================================================================
        
        self.deviceInfo.setText("Finding spectrum analyzer...")
        self.btn_findDevice.setEnabled(False)
        self.updateUi()
        try:
            self.foundSpec=self.specan.find_device()
            
        except:
            self.foundSpec=False
        
        if self.foundSpec:
            self.btn_findDevice.setEnabled(False)
            self.btn_run.setEnabled(True)
            self.deviceInfo.setText("Spectrum analyzer found!! (SH BB60C)")
            self.runInfo.setText("Ready to run test")
            self.specan.sh.configureGain('auto')
            
            #===================================================================
            # setup signal hound
            #===================================================================
            try:
                self.specan.sh.configureLevel(ref = 0, atten = 'auto')
                self.specan.sh.configureProcUnits("log")
                self.specan.sh.configureAcquisition("min-max","log-scale")
                self.specan.sh.configureCenterSpan(5500e6,1000e6)
                self.specan.sh.configureSweepCoupling(100e3,100e3,0.09,"native","no-spur-reject")
                self.specan.sh.configureGain('auto')
            except:
                print "specan setup error"
            
            
             
            
        else:
            self.deviceInfo.setText("No spectrum analyzer Found")
            self.btn_findDevice.setEnabled(True)
    
    def show_errorDialog(self,title,text,info):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Critical)
        msg.setText(text)
        msg.setInformativeText(info)
        msg.setWindowTitle(title)
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec_()
    
def atexit(app):
    app.purgeplotImageList()
        
    
def main():
    app = QApplication(sys.argv)  
    app.setStyle(QStyleFactory.create("plastique"))
    form = Application()
    form.show()
    #form.resize(400,600)
    app.exec_()
    
    atexit(form)
    sys.exit()
    
    


if __name__ == '__main__':
    main()
