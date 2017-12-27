'''
Created on Nov 10, 2017

@author: Mike
'''
from PyQt4.QtCore import *
from PyQt4.QtGui import *
from SpecTest import SpecTest

import openpyxl as pyxl
from openpyxl import load_workbook

class advSettingsDialog(QDialog):
    '''
    dialog class to setup signalhound bb60c
    '''


    def __init__(self, parent=None):

        super(advSettingsDialog,self).__init__(parent)   
        self.parent=parent
        "Initialze spectrum analyzer setting dialog box"
        self.setWindowTitle("Advanced Settings")
        self.vert = QVBoxLayout() 
        self.resize(1000,400)
        #=======================================================================
        # setup buttons
        #=======================================================================
        self.b_box = QDialogButtonBox(QDialogButtonBox.Ok  | QDialogButtonBox.Cancel)
            
        #set button behavior
        self.connect(self.b_box, SIGNAL('rejected()'),self.click_cancel)    #behavior when "cancel" is clicked
        self.connect(self.b_box, SIGNAL('accepted()'),self.click_ok)        #behavior when "ok" is clicked
        
        self.btn_add=QPushButton("Add")
        self.btn_add.setFixedHeight(50)
        self.btn_add.setFixedWidth(150)
        self.connect(self.btn_add, SIGNAL('clicked()'), self.click_add)
        
        self.btn_remove=QPushButton("Remove")
        self.btn_remove.setFixedHeight(50)
        self.btn_remove.setFixedWidth(150)
        self.connect(self.btn_remove, SIGNAL('clicked()'), self.click_remove)
        
        self.btn_import=QPushButton("Import Configuration")
        self.btn_import.setFixedHeight(50)
        self.btn_import.setFixedWidth(150)
        self.connect(self.btn_import, SIGNAL('clicked()'), self.click_import)
        
        self.btn_export=QPushButton("Export Configuration")
        self.btn_export.setFixedHeight(50)
        self.btn_export.setFixedWidth(150)
        self.connect(self.btn_export, SIGNAL('clicked()'), self.click_export)
        
        self.btn_default=QPushButton("Load Default Configuration")
        self.btn_default.setFixedHeight(50)
        self.btn_default.setFixedWidth(150)
        self.connect(self.btn_default, SIGNAL('clicked()'), self.click_default)
        
        
        #=======================================================================
        # setup table
        #=======================================================================
        self.table=QTableWidget()
        
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels(QString("Name;Center Frequency(MHz);Frequency Span(MHz);Start Freq(MHz);End Freq(MHz);RBW(MHz);Sweep Time(ms);Sweep Count;Threshold(dBm)").split(";"))
        self.table.setRowCount(0)
        self.table.cellClicked.connect(self.cellch)
        header = self.table.horizontalHeader()
        header.setResizeMode(QHeaderView.ResizeToContents)
        header.setStretchLastSection(True)
        
        self.disableTableModification=True
        
#         self.table.currentItemChanged.connect(self.setTableValues) 
       
        hbox=QHBoxLayout()
        hbox.addStretch()
        hbox.addWidget(self.btn_add)
        hbox.addWidget(self.btn_remove)
        hbox.addStretch()

        hbox2=QHBoxLayout()
        hbox2.addStretch()
        hbox2.addWidget(self.btn_import)
        hbox2.addWidget(self.btn_export)
        hbox2.addStretch()
        hbox2.addWidget(self.btn_default)
        hbox2.addStretch()
        
        
        self.vert.addLayout(hbox)
        self.vert.addWidget(self.table)
        self.vert.addLayout(hbox2)
        self.vert.addWidget(self.b_box)
        self.setLayout(self.vert)
    def cellch(self):
        print self.table.currentRow()
        print self.parent.testList[self.table.currentRow()].name
        
    def click_ok(self):
        self.setTableValues()
        self.close()
        
    def click_cancel(self):
        self.close()
    
    def click_add(self):
        self.parent.testList.append(SpecTest(parent=self.parent, plot=self.parent.plot, testNum=len(self.parent.testList), name="New Test",rbw=100e3,sweepTime=0.01,sweepNum=20,freqCenter=1e9,freqSpan=1e9,threshold=-50))
        self.updateTable()
        self.btn_remove.setEnabled(True)
    
    def click_remove(self):
        length=len(self.parent.testList)
        
        #delete list entry
        if length>1:
            row=int(self.table.currentRow())
#             del self.parent.testList[row]
            x=self.parent.testList.pop(row)
            print "deleted ", x.name
            for i in self.parent.testList:
                print i.name
            
            
            #update table
            length=len(self.parent.testList)
            if length>0:
                self.table.setCurrentCell(0, 0)
                self.updateTable()
                pass
            if length==1:
                self.btn_remove.setEnabled(False)
            
            #update test numbers for PNGs
            for i in self.parent.testList:
                if i.testNum>row:
                    i.testNum-=1;
            
    def click_import(self):
        file_choices = "Excel Workbook ( *.xlsx)"
        
        #open file 
        path = unicode(QFileDialog.getOpenFileName(self, 
                        'Open', '', 
                        file_choices))
        if path:
            try:
                wb2 = load_workbook(path)   #load openpyxl workbook
                names=wb2.get_sheet_names() #get name of data sheet
                ws=wb2[str(names[0])];      #set active worksheet to data sheet
                
                
                configVerifier=str(ws['XX1'].value)
                if configVerifier!="testConfig":
                    self.show_errorDialog("File Import Error!", "The selected file is not a valid configuration file.!", "Please import a different file.")
                    return
                
                self.parent.testList=[]
                i=0
                while ws['A'+str(i+1)].value!=None:
                    self.parent.testList.append(SpecTest(parent=self.parent, plot=self.parent.plot, testNum=len(self.parent.testList), name="New Test",rbw=100e3,sweepTime=0.01,sweepNum=20,freqCenter=1e9,freqSpan=1e9,threshold=-50))
                    self.parent.testList[i].name=str(ws['A'+str(i+1)].value)
                    self.parent.testList[i].freqCenter=int(float(ws['B'+str(i+1)].value))*1e6
                    self.parent.testList[i].freqSpan=int(float(ws['C'+str(i+1)].value))*1e6
                    self.parent.testList[i].rbw=float(ws['F'+str(i+1)].value)*1e6
                    self.parent.testList[i].sweepTime=float(ws['G'+str(i+1)].value)/1e3
                    self.parent.testList[i].sweepNum=int(float(ws['H'+str(i+1)].value))
                    self.parent.testList[i].threshold=int(float(ws['I'+str(i+1)].value))
                    i+=1
                self.updateTable()
            except:
                self.show_errorDialog("File Import Error!", "There was an error importing the Configuration File!", "Please import a different file.")
            
    
    def click_export(self):
        file_choices = "Excel Workbook ( *.xlsx)"
        path = unicode(QFileDialog.getSaveFileName(self, 
                        'Save', '', 
                        file_choices))
        
        if path:
            wb = pyxl.Workbook()
            ws = wb.create_sheet("CONFIG", 0)
            for i in range(len(self.parent.testList)):
                ws['A'+str(i+1)]=str(self.table.item(i, 0).text())
                ws['B'+str(i+1)]=str(self.table.item(i, 1).text())
                ws['C'+str(i+1)]=str(self.table.item(i, 2).text())
                ws['D'+str(i+1)]=str(self.table.item(i, 3).text())
                ws['E'+str(i+1)]=str(self.table.item(i, 4).text())
                ws['F'+str(i+1)]=str(self.table.item(i, 5).text())
                ws['G'+str(i+1)]=str(self.table.item(i, 6).text())
                ws['H'+str(i+1)]=str(self.table.item(i, 7).text())
                ws['I'+str(i+1)]=str(self.table.item(i, 8).text())
            ws['XX1']="testConfig"
            try:
                wb.save(path)
            except: 
                print "Error"
                self.show_errorDialog("File Save Error!", "Unable to save report!", "Ensure report is not open in another program.")
                  
    def click_default(self):
        self.parent.buildDefaultTests()
        self.btn_remove.setEnabled(True)
        self.updateTable()
    
    def setTableValues(self):
        
        if self.disableTableModification:
            pass
        else:
            for i in range(len(self.parent.testList)):
                self.parent.testList[i].name=str(self.table.item(i,0).text())
                self.parent.testList[i].freqCenter=int(float(self.table.item(i,1).text()))*1e6
                center=self.parent.testList[i]
                self.parent.testList[i].freqSpan=int(float(self.table.item(i,2).text()))*1e6
                span=self.parent.testList[i]
                self.parent.testList[i].rbw=float(self.table.item(i,5).text())*1e6
                self.parent.testList[i].sweepTime=float(self.table.item(i,6).text())/1e3
                self.parent.testList[i].sweepNum=int(float(self.table.item(i,7).text()))
                self.parent.testList[i].threshold=int(float(self.table.item(i,8).text())) 
            self.updateTable()
            
    def updateTable(self):
        
        rows=len(self.parent.testList)
        
        if rows>=0:
            self.table.setRowCount(rows)
            
            for i in range(rows):
                self.parent.testList[i]
                self.table.setItem(i,0,QTableWidgetItem(str(self.parent.testList[i].name)))
                center=self.parent.testList[i].freqCenter
                self.table.setItem(i,1,QTableWidgetItem(str(center/1e6)))
                span=self.parent.testList[i].freqSpan
                self.table.setItem(i,2,QTableWidgetItem(str(span/1e6)))
                
                start=int(center-span/2)
                item=QTableWidgetItem(str(start/1e6))
                item.setFlags( ~(Qt.ItemIsSelectable |  Qt.ItemIsEnabled) )
                self.table.setItem(i,3,item)
                
                end=int(center+span/2)
                item=QTableWidgetItem(str(end/1e6))
                item.setFlags( ~(Qt.ItemIsSelectable |  Qt.ItemIsEnabled) )
                self.table.setItem(i,4,item)
                
                
                self.table.setItem(i,5,QTableWidgetItem(str(self.parent.testList[i].rbw/1e6)))
                self.table.setItem(i,6,QTableWidgetItem(str(self.parent.testList[i].sweepTime*1e3)))
                self.table.setItem(i,7,QTableWidgetItem(str(self.parent.testList[i].sweepNum)))
                self.table.setItem(i,8,QTableWidgetItem(str(self.parent.testList[i].threshold)))
                
                if self.table.currentRow()==-1:
                    self.table.setCurrentCell(0, 0)
   
    def show_errorDialog(self,title,text,info):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Critical)
        msg.setText(text)
        msg.setInformativeText(info)
        msg.setWindowTitle(title)
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec_()

   
   
   
   
#EOF